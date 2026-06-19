#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lee un archivo Hojas_de_vida_indicadores_*.xlsx en el NUEVO formato dinámico
(con hoja "Lista X" donde cada fila = un indicador) y genera un CSV compatible
con website/lib/indicator_metadata.php (im_import_csv).

Uso:
  py -3 database/scripts/import_hojas_vida_dynamic_xlsx.py <ruta.xlsx> <observatory_id> [salida.csv]

Ejemplos:
  py -3 database/scripts/import_hojas_vida_dynamic_xlsx.py "C:/.../Hojas_de_vida_indicadores_Economicos_01.xlsx" 1
  py -3 database/scripts/import_hojas_vida_dynamic_xlsx.py "C:/.../Hojas_de_vida_indicadores_Social_01.xlsx" 2
  py -3 database/scripts/import_hojas_vida_dynamic_xlsx.py "C:/.../Hojas_de_vida_indicadores_Genero_01.xlsx" 5

Si no se pasa salida, escribe database/seeds/indicators_<obs>_dynamic_import.csv
"""
from __future__ import annotations
import csv, re, sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[2]

CSV_COLUMNS = [
    "id", "observatory_id", "title", "category_1", "category_2", "tags", "unit",
    "thematic_breakdown", "geographic_breakdown", "definition", "calculation_formula",
    "periodicity", "baseline_date", "delivery_form", "source", "source_link",
    "actors", "responsible_entity", "observations", "availability_status",
]

OBS_SLUG = {1: "economico", 2: "social", 3: "ambiente", 4: "cti", 5: "genero"}

def _s(v) -> str:
    if v is None:
        return ""
    t = str(v).strip()
    if t.lower() == "nan" or t.lower() == "none":
        return ""
    return t

def _one_line(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()

def find_lista_sheet(wb) -> str:
    for s in wb.sheetnames:
        norm = s.lower().strip()
        if norm.startswith("lista"):
            return s
    raise RuntimeError(f"No se encontró ninguna hoja 'Lista …' en: {wb.sheetnames}")

def parse_lista(xlsx: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    name = find_lista_sheet(wb)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    # Localizar fila de cabecera: contiene 'id'
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(_s(c).lower() == "id" for c in row[:3] if c is not None):
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError(f"No se encontró cabecera 'id' en hoja {name!r}")
    headers = [_s(c).lower() for c in rows[header_idx]]
    # Mapeo de cabecera del nuevo formato → claves internas
    H = {
        "id": "id",
        "indicador": "title",
        "categoría de primer orden": "cat1",
        "categoria de primer orden": "cat1",
        "categoría de segundo orden": "cat2",
        "categoria de segundo orden": "cat2",
        "etiquetas": "tags",
        "unidad de medida": "unit",
        "fecha de los datos": "baseline_date",
        "link archivo": "source_link",
        "hoja base dx": "hoja_dx",
        "archivo fuente": "archivo",
        "definición del indicador": "definition",
        "definicion del indicador": "definition",
        "cómo se calcula": "calculation_formula",
        "como se calcula": "calculation_formula",
        "periodicidad": "periodicity",
        "fuente": "source",
        "desagregación temática": "thematic_breakdown",
        "desagregacion tematica": "thematic_breakdown",
        "actores involucrados": "actors",
        "responsable": "responsible_entity",
        "observaciones": "observations",
        "observatorio": "observatorio",
        "medio de obtención": "delivery_form",
        "medio de obtencion": "delivery_form",
        "medios de difusión": "medios_difusion",
        "medios de difusion": "medios_difusion",
    }
    col_map = {}
    for i, h in enumerate(headers):
        key = H.get(h)
        if key:
            col_map[key] = i
    if "id" not in col_map or "title" not in col_map:
        raise RuntimeError(f"Cabeceras incompletas. Encontradas: {headers}")

    out = []
    for r in rows[header_idx + 1:]:
        if r is None:
            continue
        cid = _s(r[col_map["id"]]) if col_map["id"] < len(r) else ""
        if not cid.isdigit():
            continue
        rec = {}
        for k, idx in col_map.items():
            rec[k] = _one_line(_s(r[idx])) if idx < len(r) else ""
        out.append(rec)
    return out

def to_csv_row(rec: dict, observatory_id: int) -> dict:
    # Componer thematic_breakdown enriquecido con la hoja base DX (igual que el script anterior)
    base_thematic = rec.get("thematic_breakdown", "")
    hoja = rec.get("hoja_dx", "")
    if hoja:
        suff = f"Hoja base DX: {hoja}"
        thematic = f"{base_thematic} | {suff}".strip(" |") if base_thematic else suff
    else:
        thematic = base_thematic
    return {
        "id": rec["id"],
        "observatory_id": observatory_id,
        "title": rec.get("title", ""),
        "category_1": rec.get("cat1", ""),
        "category_2": rec.get("cat2", ""),
        "tags": rec.get("tags", ""),
        "unit": rec.get("unit", ""),
        "thematic_breakdown": thematic,
        "geographic_breakdown": "Departamento / Municipio",
        "definition": rec.get("definition", ""),
        "calculation_formula": rec.get("calculation_formula", ""),
        "periodicity": rec.get("periodicity", ""),
        "baseline_date": rec.get("baseline_date", ""),
        "delivery_form": rec.get("delivery_form", ""),
        "source": rec.get("source", ""),
        "source_link": rec.get("source_link", ""),
        "actors": rec.get("actors", ""),
        "responsible_entity": rec.get("responsible_entity", ""),
        "observations": rec.get("observations", ""),
        "availability_status": "DISPONIBLE",
    }

def write_csv(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)

def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__)
        return 2
    xlsx = sys.argv[1]
    try:
        obs_id = int(sys.argv[2])
    except ValueError:
        print("observatory_id debe ser un número entero (1=Eco, 2=Soc, 5=Género).", file=sys.stderr)
        return 2
    if not Path(xlsx).is_file():
        print(f"No existe el archivo: {xlsx}", file=sys.stderr)
        return 1
    if len(sys.argv) > 3:
        out = Path(sys.argv[3])
    else:
        slug = OBS_SLUG.get(obs_id, f"obs{obs_id}")
        out = ROOT / "database/seeds" / f"indicators_{slug}_dynamic_import.csv"

    raw = parse_lista(xlsx)
    if not raw:
        print("No se extrajeron filas de la hoja Lista.", file=sys.stderr)
        return 1
    csv_rows = [to_csv_row(r, obs_id) for r in raw]
    write_csv(csv_rows, out)
    ids = sorted(int(r["id"]) for r in csv_rows)
    print(f"Extraídas {len(csv_rows)} filas -> {out}")
    print(f"Rango códigos: {ids[0]} … {ids[-1]}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
