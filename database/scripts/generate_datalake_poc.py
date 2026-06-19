#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera el datalake de 3 indicadores como prueba de concepto:
  - 1000  Analfabetismo (Económico)        → Pob Multidimensional
  - 2000  Estadísticas demográficas (Social) → Demografía
  - 5000  Estadísticas demográficas con enfoque de género (Género) → Demografía

Cada datalake genera la carpeta website/indicador/{ID}/ con:
  indicador.info, 1.info, 1.csv, 2.info, 2.csv, display.js
"""
from __future__ import annotations
import csv, sys, os
from pathlib import Path
from collections import defaultdict
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[2]
WEBSITE_INDICADOR = ROOT / "website" / "indicador"
HV_BASE = Path(r"C:\Users\cesar\Documents\Gobernación\Observatorio\Hoja de vida indicadores")

# ── Helpers ────────────────────────────────────────────────────────────────
def write_info(path: Path, fields: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for k, v in fields.items():
            f.write(f"{k}:{v}\n")

def write_csv(path: Path, rows: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerows(rows)

def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        f.write(content)

def parse_year(v):
    if v is None: return None
    s = str(v)
    digits = ''.join(c for c in s if c.isdigit())[:4]
    try:
        y = int(digits)
        return y if 1990 <= y <= 2030 else None
    except: return None

# ── 1000 Analfabetismo (Económico) ─────────────────────────────────────────
def build_1000():
    src = HV_BASE / "HV indicadores economico" / "BASE DX OBS ECONÓMICO.xlsx"
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb["Pob Multidimensional"]
    rows_total = []  # [year, total]
    rows_split = []  # [year, cabeceras, centros poblados/rural]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != "Boyacá" or row[1] != "Analfabetismo":
            continue
        y = parse_year(row[2])
        if y is None: continue
        total = row[4]
        cab   = row[5]
        rur   = row[6]
        rows_total.append([y, total])
        rows_split.append([y, cab, rur])
    wb.close()
    rows_total.sort(key=lambda x: x[0])
    rows_split.sort(key=lambda x: x[0])

    folder = WEBSITE_INDICADOR / "1000"
    write_info(folder / "indicador.info", {
        "Categoría": "Pobreza",
        "Descripción": "Porcentaje de personas de 15 años o más que no saben leer ni escribir, según la metodología del Índice de Pobreza Multidimensional (IPM) del DANE. Se desagrega entre cabeceras y centros poblados / rural disperso.",
        "Titulo": "Analfabetismo",
        "Subcategoría": "Pobreza Multidimensional",
        "Etiquetas": "ND",
        "Fuentes": "https://www.dane.gov.co/index.php/estadisticas-por-tema/pobreza-y-condiciones-de-vida/pobreza-multidimensional",
    })
    # Chart 1: serie temporal total
    write_info(folder / "1.info", {
        "Titulo": "Tasa de analfabetismo en Boyacá",
        "Descripción": "Evolución del porcentaje de personas que no saben leer ni escribir en el departamento.",
        "Vertical": "Porcentaje",
        "Horizontal": "Año",
    })
    write_csv(folder / "1.csv", [["Año", "Boyacá"]] + rows_total)
    # Chart 2: comparación urbano/rural
    write_info(folder / "2.info", {
        "Titulo": "Analfabetismo por área geográfica",
        "Descripción": "Comparación de la tasa de analfabetismo entre cabeceras municipales y zonas rurales / centros poblados.",
        "Vertical": "Porcentaje",
        "Horizontal": "Año",
    })
    write_csv(folder / "2.csv", [["Año", "Cabeceras", "Centros poblados y rural disperso"]] + rows_split)
    # display.js — 2 charts línea
    write_text(folder / "display.js", DISPLAY_TWO_LINES)
    print(f"[1000] OK · 2 charts · {len(rows_total)} años")

# ── 2000 Demografía (Social) ───────────────────────────────────────────────
def build_2000():
    src = HV_BASE / "HV indicadores Social" / "BASE DX OBS SOCIAL - DEMOGRAFÍA.xlsx"
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb["Demografía"]
    by_year_total = defaultdict(int)
    by_year_sex   = defaultdict(lambda: {"Hombres": 0, "Mujeres": 0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        # cols: DP, DPNOM, MPIO, DPMP, AÑO, ÁREA, Genero, SEXO, Edad, TOTAL...
        if row[1] != "Boyacá": continue
        if row[5] != "Total": continue       # Solo total (no duplicar urbano+rural)
        y = parse_year(row[4])
        if y is None: continue
        try:
            tot = int(row[9]) if row[9] is not None else 0
        except: continue
        sexo = row[7]  # 'Hombres' o 'Mujeres'
        by_year_total[y] += tot
        if sexo in ("Hombres", "Mujeres"):
            by_year_sex[y][sexo] += tot
    wb.close()
    years = sorted(by_year_total.keys())
    rows_total = [[y, by_year_total[y]] for y in years]
    rows_sex   = [[y, by_year_sex[y]["Hombres"], by_year_sex[y]["Mujeres"]] for y in years]

    folder = WEBSITE_INDICADOR / "2000"
    write_info(folder / "indicador.info", {
        "Categoría": "Demografía",
        "Descripción": "Estadísticas y proyecciones de población del departamento de Boyacá según las proyecciones del DANE, con desagregación por sexo, edad, municipio y área geográfica.",
        "Titulo": "Estadísticas y proyecciones demográficas",
        "Subcategoría": "Población",
        "Etiquetas": "ND",
        "Fuentes": "https://www.dane.gov.co/index.php/estadisticas-por-tema/demografia-y-poblacion",
    })
    write_info(folder / "1.info", {
        "Titulo": "Población total de Boyacá",
        "Descripción": "Proyección anual de la población total del departamento.",
        "Vertical": "Habitantes",
        "Horizontal": "Año",
    })
    write_csv(folder / "1.csv", [["Año", "Población"]] + rows_total)
    write_info(folder / "2.info", {
        "Titulo": "Población por sexo",
        "Descripción": "Comparación de la población masculina y femenina del departamento, año a año.",
        "Vertical": "Habitantes",
        "Horizontal": "Año",
    })
    write_csv(folder / "2.csv", [["Año", "Hombres", "Mujeres"]] + rows_sex)
    write_text(folder / "display.js", DISPLAY_LINE_PLUS_BARS)
    print(f"[2000] OK · 2 charts · {len(years)} años")

# ── 5000 Demografía con enfoque género (Género) ────────────────────────────
def build_5000():
    src = HV_BASE / "HV indicadores Asuntos de Género" / "BASE DX OBS ASUNTOS DE GÉNERO.xlsx"
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb["Demografía"]
    by_year_total = defaultdict(int)                            # Población femenina total
    by_year_area  = defaultdict(lambda: {"Cabecera": 0, "Rural": 0})
    by_ciclo      = defaultdict(int)                            # Por ciclo de vida (último año)
    last_year_per_municipio = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != "Boyacá": continue
        y = parse_year(row[4])
        if y is None: continue
        area = (row[5] or "").strip()
        try:
            tot = int(row[9]) if row[9] is not None else 0
        except: continue
        ciclo = row[10] if len(row) > 10 else None
        if area == "Total":
            by_year_total[y] += tot
            if y > last_year_per_municipio:
                last_year_per_municipio = y
        elif area == "Cabecera Municipal":
            by_year_area[y]["Cabecera"] += tot
        elif area == "Centros Poblados y Rural Disperso":
            by_year_area[y]["Rural"] += tot
    # Segunda pasada para ciclo de vida del último año
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] != "Boyacá": continue
        y = parse_year(row[4])
        if y != last_year_per_municipio: continue
        if (row[5] or "").strip() != "Total": continue
        try:
            tot = int(row[9]) if row[9] is not None else 0
        except: continue
        ciclo = row[10] if len(row) > 10 else None
        if ciclo:
            by_ciclo[str(ciclo).strip()] += tot
    wb.close()

    years = sorted(by_year_total.keys())
    rows_total = [[y, by_year_total[y]] for y in years]
    rows_area  = [[y, by_year_area[y]["Cabecera"], by_year_area[y]["Rural"]] for y in years]
    # Orden de ciclos de vida típico
    ciclo_order = ["Primera Infancia (0 a 5 años)", "Infancia (6 a 11 años)", "Adolescencia (12 a 17 años)",
                   "Juventud (18 a 28 años)", "Adultez (29 a 59 años)", "Vejez (60 años o más)"]
    rows_ciclo = [[c, by_ciclo[c]] for c in ciclo_order if c in by_ciclo]
    # incluir cualquier otro ciclo no listado al final
    for k, v in by_ciclo.items():
        if k not in ciclo_order:
            rows_ciclo.append([k, v])

    folder = WEBSITE_INDICADOR / "5000"
    write_info(folder / "indicador.info", {
        "Categoría": "Demografía y población",
        "Descripción": "Estadísticas demográficas de la población femenina del departamento de Boyacá. Incluye proyección anual, distribución por área geográfica y por ciclo de vida.",
        "Titulo": "Estadísticas demográficas con enfoque de género",
        "Subcategoría": "Población femenina",
        "Etiquetas": "ND",
        "Fuentes": "https://www.dane.gov.co/index.php/estadisticas-por-tema/demografia-y-poblacion",
    })
    write_info(folder / "1.info", {
        "Titulo": "Población femenina total en Boyacá",
        "Descripción": "Proyección anual de la población femenina del departamento.",
        "Vertical": "Habitantes",
        "Horizontal": "Año",
    })
    write_csv(folder / "1.csv", [["Año", "Mujeres"]] + rows_total)
    write_info(folder / "2.info", {
        "Titulo": "Población femenina por área geográfica",
        "Descripción": "Distribución de la población femenina entre cabeceras municipales y zonas rurales / centros poblados.",
        "Vertical": "Habitantes",
        "Horizontal": "Año",
    })
    write_csv(folder / "2.csv", [["Año", "Cabecera Municipal", "Rural / Centros Poblados"]] + rows_area)
    write_info(folder / "3.info", {
        "Titulo": f"Población femenina por ciclo de vida ({last_year_per_municipio})",
        "Descripción": f"Distribución de la población femenina por ciclo de vida en el último año disponible ({last_year_per_municipio}).",
        "Vertical": "Habitantes",
        "Horizontal": "Ciclo de vida",
    })
    write_csv(folder / "3.csv", [["Ciclo", "Mujeres"]] + rows_ciclo)
    write_text(folder / "display.js", DISPLAY_LINE_LINE_BARS)
    print(f"[5000] OK · 3 charts · {len(years)} años · {len(rows_ciclo)} ciclos")

# ── Plantillas display.js ──────────────────────────────────────────────────
DISPLAY_TWO_LINES = """class Chart1 extends AbstractChart{
    format(){
        var yearFormatter = new google.visualization.NumberFormat({pattern: Patterns.year});
        yearFormatter.format(this._data, 0);
    }
    getOptions(info){
        return {
            hAxis: {title: info['horizontal'], format: Patterns.year},
            vAxis: {title: info['vertical']},
            curveType: 'function',
            pointSize: 6
        };
    }
    getType(div){ return new google.visualization.LineChart(div); }
}

class Chart2 extends AbstractChart{
    format(){
        var yearFormatter = new google.visualization.NumberFormat({pattern: Patterns.year});
        yearFormatter.format(this._data, 0);
    }
    getOptions(info){
        return {
            hAxis: {title: info['horizontal'], format: Patterns.year},
            vAxis: {title: info['vertical']},
            curveType: 'function',
            pointSize: 6
        };
    }
    getType(div){ return new google.visualization.LineChart(div); }
}

class Display extends AbstractDisplay{
    constructor(){ super('corechart',[Chart1,Chart2]); }
}
"""

DISPLAY_LINE_PLUS_BARS = """class Chart1 extends AbstractChart{
    format(){
        var yearFormatter = new google.visualization.NumberFormat({pattern: Patterns.year});
        yearFormatter.format(this._data, 0);
    }
    getOptions(info){
        return {
            hAxis: {title: info['horizontal'], format: Patterns.year},
            vAxis: {title: info['vertical']},
            curveType: 'function',
            pointSize: 6
        };
    }
    getType(div){ return new google.visualization.LineChart(div); }
}

class Chart2 extends AbstractChart{
    format(){
        var yearFormatter = new google.visualization.NumberFormat({pattern: Patterns.year});
        yearFormatter.format(this._data, 0);
    }
    getOptions(info){
        return {
            hAxis: {title: info['horizontal'], format: Patterns.year},
            vAxis: {title: info['vertical']},
            isStacked: false,
            bar: {groupWidth: '70%'}
        };
    }
    getType(div){ return new google.visualization.ColumnChart(div); }
}

class Display extends AbstractDisplay{
    constructor(){ super('corechart',[Chart1,Chart2]); }
}
"""

DISPLAY_LINE_LINE_BARS = """class Chart1 extends AbstractChart{
    format(){
        var yearFormatter = new google.visualization.NumberFormat({pattern: Patterns.year});
        yearFormatter.format(this._data, 0);
    }
    getOptions(info){
        return {
            hAxis: {title: info['horizontal'], format: Patterns.year},
            vAxis: {title: info['vertical']},
            curveType: 'function',
            pointSize: 6
        };
    }
    getType(div){ return new google.visualization.LineChart(div); }
}

class Chart2 extends AbstractChart{
    format(){
        var yearFormatter = new google.visualization.NumberFormat({pattern: Patterns.year});
        yearFormatter.format(this._data, 0);
    }
    getOptions(info){
        return {
            hAxis: {title: info['horizontal'], format: Patterns.year},
            vAxis: {title: info['vertical']},
            curveType: 'function',
            pointSize: 6
        };
    }
    getType(div){ return new google.visualization.LineChart(div); }
}

class Chart3 extends AbstractChart{
    getOptions(info){
        return {
            hAxis: {title: info['horizontal']},
            vAxis: {title: info['vertical']},
            bar: {groupWidth: '70%'},
            legend: {position: 'none'}
        };
    }
    getType(div){ return new google.visualization.ColumnChart(div); }
}

class Display extends AbstractDisplay{
    constructor(){ super('corechart',[Chart1,Chart2,Chart3]); }
}
"""

if __name__ == "__main__":
    build_1000()
    build_2000()
    build_5000()
    print("\n✅ Datalake PoC generado en website/indicador/{1000,2000,5000}")
