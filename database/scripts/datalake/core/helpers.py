"""
Helpers comunes para la generación de datalake automatizada.
"""
from __future__ import annotations
import csv, re
from pathlib import Path
import openpyxl

# ── Rutas absolutas (estables) ─────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[4]  # Observatorio2026/
WEBSITE_INDICADOR = ROOT / "website" / "indicador"
HV_BASE = Path(r"C:\Users\cesar\Documents\Gobernación\Observatorio\Hoja de vida indicadores")

# Mapping observatorio → carpeta y archivo de origen
SOURCE_FILES = {
    1: HV_BASE / "HV indicadores economico"        / "BASE DX OBS ECONÓMICO.xlsx",
    2: HV_BASE / "HV indicadores Social"            / "BASE DX OBS SOCIAL - DEMOGRAFÍA.xlsx",   # default; algunas hojas viven en otros archivos
    5: HV_BASE / "HV indicadores Asuntos de Género" / "BASE DX OBS ASUNTOS DE GÉNERO.xlsx",
}

# Para Social, una misma hoja puede vivir en varios archivos (las líneas).
# Búsqueda por nombre de hoja entre todos los archivos del observatorio.
SOCIAL_FILES_DIR = HV_BASE / "HV indicadores Social"
SOCIAL_FILES = list(SOCIAL_FILES_DIR.glob("BASE DX OBS SOCIAL - *.xlsx")) if SOCIAL_FILES_DIR.exists() else []

# ── Parsers ─────────────────────────────────────────────────────────────────
YEAR_PAT = re.compile(r'^\s*(a[ñn]o|year|fecha)\s*$', re.I)

def parse_year(v):
    if v is None:
        return None
    s = str(v)
    digits = ''.join(c for c in s if c.isdigit())[:4]
    try:
        y = int(digits)
        return y if 1990 <= y <= 2030 else None
    except (ValueError, TypeError):
        return None

def to_number(v):
    """Convierte valor a número, o None si no es convertible."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(',', '.')
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return None

def normalize_str(s):
    """Lowercase + sin tildes, para comparaciones robustas."""
    if s is None:
        return ''
    s = str(s).strip().lower()
    return s.translate(str.maketrans('áéíóúñ', 'aeiouñ')).replace('ñ', 'n')

# ── Carga de Excel ─────────────────────────────────────────────────────────
def find_workbook_for_sheet(observatory_id: int, sheet_name: str) -> Path | None:
    """Localiza el archivo .xlsx que contiene la hoja indicada para un observatorio."""
    if observatory_id == 2:
        candidates = SOCIAL_FILES
    else:
        candidates = [SOURCE_FILES.get(observatory_id)]
    target = normalize_str(sheet_name)
    for path in candidates:
        if path is None or not path.exists():
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            for s in wb.sheetnames:
                if normalize_str(s) == target:
                    wb.close()
                    return path
            wb.close()
        except Exception:
            continue
    return None

def load_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[tuple]]:
    """Carga una hoja y devuelve (headers, rows)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # Buscar nombre exacto (insensible a mayúsculas/tildes)
    target = normalize_str(sheet_name)
    actual = None
    for s in wb.sheetnames:
        if normalize_str(s) == target:
            actual = s
            break
    if actual is None:
        wb.close()
        raise ValueError(f"Hoja {sheet_name!r} no existe en {path.name}")
    ws = wb[actual]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else '' for c in next(rows_iter, ())]
    data_rows = []
    cnt = 0
    for row in rows_iter:
        cnt += 1
        if cnt > 200000:  # límite de seguridad para hojas masivas
            break
        if any(c is not None for c in row):
            data_rows.append(row)
    wb.close()
    return headers, data_rows

def find_col(headers: list[str], *candidates: str) -> int | None:
    """Busca el índice de la primera columna que coincida (insensible)."""
    norm = [normalize_str(h) for h in headers]
    for cand in candidates:
        c = normalize_str(cand)
        for i, h in enumerate(norm):
            if h == c or c in h:
                return i
    return None

# ── Escritores ──────────────────────────────────────────────────────────────
def write_info(path: Path, fields: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for k, v in fields.items():
            f.write(f"{k}:{v}\n")

def write_csv(path: Path, rows: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerows(rows)

def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        f.write(content)

def write_indicator_files(indicator_id: int, indicator_info: dict, charts: list[dict]) -> Path:
    """
    charts: [{'titulo','descripcion','vertical','horizontal','tipo','rows':[[...],[...]],'display_class':'Line'|'Column'|'Map'}, ...]
    """
    folder = WEBSITE_INDICADOR / str(indicator_id)
    folder.mkdir(parents=True, exist_ok=True)

    # Limpiar archivos numerados antiguos (1.info, 1.csv, 2.info, ...) antes de escribir los nuevos
    # para evitar que sobren visualizaciones de generaciones previas.
    for f in folder.iterdir():
        nm = f.name
        if nm == "indicador.info" or nm == "display.js":
            continue
        # 1.info, 12.csv, etc. — patron N.info / N.csv
        if "." in nm:
            stem, ext = nm.rsplit(".", 1)
            if stem.isdigit() and ext in ("info", "csv"):
                try: f.unlink()
                except Exception: pass

    # indicador.info
    write_info(folder / "indicador.info", indicator_info)

    # cada chart
    display_classes = []
    for i, c in enumerate(charts, start=1):
        info_fields = {
            "Titulo": c.get("titulo", f"Visualización {i}"),
            "Descripción": c.get("descripcion", ""),
            "Vertical": c.get("vertical", ""),
            "Horizontal": c.get("horizontal", ""),
        }
        if c.get("tipo"):
            info_fields["Tipo"] = c["tipo"]
        if c.get("fuentes"):
            info_fields["Fuentes"] = c["fuentes"]
        write_info(folder / f"{i}.info", info_fields)
        write_csv(folder / f"{i}.csv", c.get("rows", []))
        display_classes.append(c.get("display_class", "Line"))

    # display.js
    from .chart_templates import build_display_js
    write_text(folder / "display.js", build_display_js(display_classes))
    return folder
